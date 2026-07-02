"""Read the HR contacts Excel sheet, and write a starter template."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Recognized columns. `title` = the recipient's job title (e.g. "Head HR");
# `role` = the job the candidate is targeting; `notes` = extra personalization.
COLUMNS = ["name", "email", "company", "title", "role", "notes"]


@dataclass
class Contact:
    name: str
    email: str
    company: str
    title: str = ""
    role: str = ""
    notes: str = ""


def write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "contacts"
    ws.append(COLUMNS)
    ws.append(
        ["Jane Doe", "jane@acme.com", "Acme Corp", "Head of HR",
         "Senior Backend Engineer",
         "Hiring backend engineers; posted about scaling their platform"]
    )
    wb.save(path)


def read_contacts(path: Path) -> list[Contact]:
    if not path.exists():
        raise FileNotFoundError(f"Contacts sheet not found: {path}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows)]
    idx = {col: header.index(col) for col in COLUMNS if col in header}
    if "email" not in idx or "name" not in idx:
        raise ValueError(
            f"Sheet must have at least 'name' and 'email' columns. Found: {header}"
        )

    out: list[Contact] = []
    for row in rows:
        if row is None:
            continue

        def get(col, _row=row):
            # Rows can be shorter than the header (ragged sheets) — guard the index.
            if col not in idx or idx[col] >= len(_row) or _row[idx[col]] is None:
                return ""
            return str(_row[idx[col]]).strip()

        email = get("email")
        if not email or "@" not in email:
            continue
        out.append(
            Contact(
                name=get("name"),
                email=email,
                company=get("company"),
                title=get("title"),
                role=get("role"),
                notes=get("notes"),
            )
        )
    return out
